import numpy as np
import pandas as pd
import time
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, Series

from deps.gdrive import GoogleDriveOperations
from deps.tech_team_database.dependencies.DatabaseSQLOperations import TpSQL


class Dashboard:
    """
    Looped retrieval, update and upload of an internal tree request
    dashboard. Update Time Independent but will be looped by main hourly.

    INCLUDES: 
    schoolcode
    # Trees Sold
    % Tree Goal Reached
    # Saplings Over last 7 days
        

    Uploaded via existing gdrive wrapper.

    """    
    
    def __init__(self):
        """ Loop through full processes once called. """
        
        self.schoolcodes = pd.read_csv("schoolcodes.csv")["schoolid"].values
        self.g = GoogleDriveOperations.GDrive() # Google Drive handle

        self.tpsql = TpSQL() # database handle
        orders = self.retrieve_data()
               
        data = self.update_data(orders)
        self.upload_data(data, orders)
            
    def retrieve_data(self):
        """ Making call to tree_order table... nothing else needed. """
        return self.tpsql.getTable("tree_order")
    
    def update_data(self, orders):
        """ Initializing base ids & calling all of the data update functions.
        
        Args:
            orders (df): current tree_order table
        """
        
        data = pd.DataFrame(index=self.schoolcodes)
        data["sapling count"] = self.get_sapling_counts(orders, self.schoolcodes)
        data["weekly sapling count"] = self.get_recent_sapling_counts(orders, self.schoolcodes, 7)
                
        return data
    
    
    def upload_data(self, data, orders):
        
        # First, write xlsx workbook (is just here locally in prod-analytics)
        wb = Workbook()

        # Getting sheets & loop through dataframe assigning vals
        ws1 = wb.create_sheet("Total Counts",0)
        ws2 = wb.create_sheet("Progress Charts",1)
        ws3 = wb.create_sheet("Unaggregated Data",2)
        
        # Sheet 1: Clean Count Data
        for r in dataframe_to_rows(data, index=True):
            ws1.append(r)

        # Sheet 3: Unaggregated Data to go into chart
        for r in dataframe_to_rows(orders, index=True):
            ws3.append(r)
     
        wb.save("Tree Request Dash.xlsx")
        
        # Uploading using drive
        fileID = "1ok7JfQjYaCu4KFusElqI-HXbq3aAj6Gn" # gDrive file ID for the sheet
        self.g.updateFile("Tree Request Dash.xlsx",fileID)
        print("https://drive.google.com/file/d/" + str(fileID) + "/view?usp=sharing")

    #########################
    # SPECIFIC UPDATE FUNCS #
    #########################

            
    def get_sapling_counts(self, orders, schoolcodes):
        """ Getting list of sapling counts for the entire event cycle.

        Args:
            orders (df): current tree_order table
            schoolcodes (list): list of school ids
        """
        sapling_counts = [] 
        for code in schoolcodes:
            
            rows = orders[orders["schoolid"] == code]
            sapling_counts.append(int(sum(rows["number"].values)))
            
        return sapling_counts
    
    def get_recent_sapling_counts(self, orders, schoolcodes, days):
        """ Getting list of sapling counts for the entire event cycle.

        Args:
            orders (df): current tree_order table
            schoolcodes (list): list of school ids
            days (int): cutoff for recent count
        """
        sapling_counts = [] 
        unix_cutoff = time.time() - 86400 * days # Current time - Number of seconds in a day * Number of days
        
        for code in schoolcodes:
            
            rows = orders[orders["schoolid"] == code]
            rows = rows[rows["submit_time"] >= unix_cutoff]
            sapling_counts.append(int(sum(rows["number"].values)))
    
        return sapling_counts
    
# Call
dash = Dashboard()